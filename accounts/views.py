import shutil
from datetime import timedelta
from decimal import Decimal

from django.conf import settings
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError
from django.views.decorators.csrf import ensure_csrf_cookie, csrf_exempt
from django.db.models import Case, Count, DecimalField, F, Q, Sum, Value, When
from django.db.models.functions import Coalesce, TruncDate, TruncMonth
from django.http import FileResponse, HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from rest_framework import status, viewsets
from rest_framework.authentication import SessionAuthentication
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.permissions import AllowAny, IsAdminUser, IsAuthenticated
from rest_framework.response import Response

from .models import Account, Category, Company, DuePayment, Note, Party, Transaction, Stock, Sale, UserProfile
from .serializers import (
    AccountSerializer,
    CategorySerializer,
    DuePaymentSerializer,
    NoteSerializer,
    PartySerializer,
    TransactionSerializer,
    StockSerializer,
    SaleSerializer,
)


def money(value):
    return Decimal(value or 0)


def filtered_transactions(request):
    qs = Transaction.objects.select_related("category", "account", "party")
    if request.user.is_authenticated:
        try:
            qs = qs.filter(company=request.user.profile.company)
        except AttributeError:
            pass
    start = request.query_params.get("start")
    end = request.query_params.get("end")
    category = request.query_params.get("category")
    method = request.query_params.get("payment_method")
    min_amount = request.query_params.get("min_amount")
    max_amount = request.query_params.get("max_amount")
    keyword = request.query_params.get("keyword")
    tx_type = request.query_params.get("transaction_type")

    if start:
        qs = qs.filter(date__gte=start)
    if end:
        qs = qs.filter(date__lte=end)
    if category:
        qs = qs.filter(category_id=category)
    if method:
        qs = qs.filter(payment_method=method)
    if min_amount:
        qs = qs.filter(amount__gte=min_amount)
    if max_amount:
        qs = qs.filter(amount__lte=max_amount)
    if tx_type:
        qs = qs.filter(transaction_type=tx_type)
    if keyword:
        qs = qs.filter(
            Q(title__icontains=keyword)
            | Q(notes__icontains=keyword)
            | Q(reference_number__icontains=keyword)
            | Q(party__name__icontains=keyword)
            | Q(category__name__icontains=keyword)
        )
    return qs


class AccountViewSet(viewsets.ModelViewSet):
    serializer_class = AccountSerializer
    queryset = Account.objects.all()

    def get_queryset(self):
        company = self.request.user.profile.company if self.request.user.is_authenticated and hasattr(self.request.user, 'profile') else None
        delta = Sum(
            Case(
                When(transactions__transaction_type="income", then=F("transactions__amount")),
                When(transactions__transaction_type="expense", then=-F("transactions__amount")),
                default=Value(0),
                output_field=DecimalField(max_digits=14, decimal_places=2),
            )
        )
        return Account.objects.filter(company=company).annotate(
            current_balance=F("opening_balance") + Coalesce(delta, Value(0, output_field=DecimalField(max_digits=14, decimal_places=2)))
        )

    def perform_create(self, serializer):
        company = self.request.user.profile.company if self.request.user.is_authenticated and hasattr(self.request.user, 'profile') else None
        serializer.save(company=company)


class CategoryViewSet(viewsets.ModelViewSet):
    serializer_class = CategorySerializer
    queryset = Category.objects.all()

    def get_queryset(self):
        company = self.request.user.profile.company if self.request.user.is_authenticated and hasattr(self.request.user, 'profile') else None
        return Category.objects.filter(company=company).order_by("category_type", "name")

    def perform_create(self, serializer):
        company = self.request.user.profile.company if self.request.user.is_authenticated and hasattr(self.request.user, 'profile') else None
        serializer.save(company=company)


class PartyViewSet(viewsets.ModelViewSet):
    serializer_class = PartySerializer
    queryset = Party.objects.all()

    def get_queryset(self):
        company = self.request.user.profile.company if self.request.user.is_authenticated and hasattr(self.request.user, 'profile') else None
        return Party.objects.filter(company=company).order_by("name")

    def perform_create(self, serializer):
        company = self.request.user.profile.company if self.request.user.is_authenticated and hasattr(self.request.user, 'profile') else None
        serializer.save(company=company)


class TransactionViewSet(viewsets.ModelViewSet):
    serializer_class = TransactionSerializer
    queryset = Transaction.objects.select_related("category", "account", "party")

    def get_queryset(self):
        return filtered_transactions(self.request)

    def list(self, request, *args, **kwargs):
        serializer = self.get_serializer(self.get_queryset()[:200], many=True)
        return Response(serializer.data)

    def perform_create(self, serializer):
        company = self.request.user.profile.company if self.request.user.is_authenticated and hasattr(self.request.user, 'profile') else None
        serializer.save(company=company)


class DuePaymentViewSet(viewsets.ModelViewSet):
    serializer_class = DuePaymentSerializer
    queryset = DuePayment.objects.select_related("party")

    def get_queryset(self):
        company = self.request.user.profile.company if self.request.user.is_authenticated and hasattr(self.request.user, 'profile') else None
        return DuePayment.objects.filter(company=company).select_related("party")

    def perform_create(self, serializer):
        company = self.request.user.profile.company if self.request.user.is_authenticated and hasattr(self.request.user, 'profile') else None
        serializer.save(company=company)


class NoteViewSet(viewsets.ModelViewSet):
    serializer_class = NoteSerializer
    queryset = Note.objects.all()

    def get_queryset(self):
        company = self.request.user.profile.company if self.request.user.is_authenticated and hasattr(self.request.user, 'profile') else None
        return Note.objects.filter(company=company)

    def perform_create(self, serializer):
        company = self.request.user.profile.company if self.request.user.is_authenticated and hasattr(self.request.user, 'profile') else None
        serializer.save(company=company)


@api_view(["GET"])
@permission_classes([AllowAny])
@ensure_csrf_cookie
def csrf_view(request):
    from django.middleware.csrf import get_token
    return Response({
        "detail": "CSRF cookie set.",
        "csrfToken": get_token(request)
    })


@api_view(["POST"])
@permission_classes([AllowAny])
def register_view(request):
    business_name = request.data.get("business_name")
    owner_name = request.data.get("owner_name")
    email = request.data.get("email")
    phone = request.data.get("phone")
    password = request.data.get("password")
    logo_base64 = request.data.get("logo_base64")

    if not business_name or not owner_name or not email or not phone or not password:
        return Response({"detail": "All fields are required."}, status=status.HTTP_400_BAD_REQUEST)

    from django.contrib.auth.models import User
    if User.objects.filter(username=email).exists() or User.objects.filter(email=email).exists():
        return Response({"detail": "This email is already registered."}, status=status.HTTP_400_BAD_REQUEST)

    from django.db import transaction
    if Company.objects.filter(name=business_name).exists():
        return Response({"detail": "This business name is already registered."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        with transaction.atomic():
            company = Company.objects.create(name=business_name, logo_base64=logo_base64)
            user = User.objects.create(
                username=email,
                email=email,
                first_name=owner_name,
                is_staff=True,
                is_superuser=True
            )
            user.set_password(password)
            user.save()

            UserProfile.objects.create(
                user=user,
                business_name=business_name,
                owner_name=owner_name,
                phone=phone,
                company=company,
                role='admin'
            )

        return Response({
            "detail": "Account created successfully! Please login.",
            "username": user.username,
            "company_name": company.name
        })
    except Exception as e:
        return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
@permission_classes([AllowAny])
def login_view(request):
    username_or_business = request.data.get("username")
    password = request.data.get("password")

    user = authenticate(username=username_or_business, password=password)

    if not user and username_or_business:
        # Try to find a user by their associated business name or company name
        profile = UserProfile.objects.filter(company__name__iexact=username_or_business).first()
        if not profile:
            profile = UserProfile.objects.filter(business_name__iexact=username_or_business).first()
        
        if profile:
            user = authenticate(username=profile.user.username, password=password)

    if not user or not user.is_staff:
        return Response({"detail": "Invalid credentials."}, status=status.HTTP_400_BAD_REQUEST)

    # Superadmin blocking & trial period verification checks
    try:
        profile = None
        if hasattr(user, 'profile'):
            profile = user.profile
        else:
            profile, created = UserProfile.objects.get_or_create(user=user)
            if created:
                from datetime import date, timedelta
                profile.trial_expiry_date = date.today() + timedelta(days=90)
                profile.save()

        if profile:
            if profile.is_blocked:
                return Response({"detail": "Your account has been blocked by the administrator."}, status=status.HTTP_403_FORBIDDEN)
            
            if not profile.is_portal_admin:
                company = profile.company
                if company and not company.is_upgraded:
                    from django.utils import timezone
                    from datetime import timedelta
                    expiry = profile.trial_expiry_date or (profile.created_at.date() + timedelta(days=90))
                    if timezone.now().date() > expiry:
                        return Response({
                            "detail": "Your 3-month trial period has expired. Please contact NMZ Associates to upgrade your account."
                        }, status=status.HTTP_403_FORBIDDEN)
    except Exception:
        pass
    
    if not user.is_superuser:
        user.is_superuser = True
        user.save()
        
    login(request, user)
    
    company_name = "Default Business"
    role = "staff"
    owner_name = user.first_name or user.username
    is_portal_admin = False
    try:
        profile = None
        if hasattr(user, 'profile'):
            profile = user.profile
        else:
            profile, created = UserProfile.objects.get_or_create(user=user)
            if created:
                from datetime import date, timedelta
                profile.trial_expiry_date = date.today() + timedelta(days=90)
                profile.save()

        if profile:
            # Automatically designate the default 'admin' and 'moosa' accounts as portal admins (case-insensitive)
            if user.username.lower() in ['admin', 'moosa'] and not profile.is_portal_admin:
                profile.is_portal_admin = True
                profile.save()
                
            company_name = profile.company.name if profile.company else "Default Business"
            role = profile.role
            owner_name = profile.owner_name or user.first_name or user.username
            is_portal_admin = profile.is_portal_admin
    except Exception:
        pass

    company_logo = None
    try:
        if hasattr(user, 'profile') and user.profile.company:
            company_logo = user.profile.company.logo_base64
    except Exception:
        pass

    return Response({
        "username": user.username,
        "owner_name": owner_name,
        "is_staff": user.is_staff,
        "company_name": company_name,
        "role": role,
        "is_portal_admin": is_portal_admin,
        "company_logo": company_logo
    })



class CsrfExemptSessionAuthentication(SessionAuthentication):
    def enforce_csrf(self, request):
        return


@api_view(["POST"])
@authentication_classes([CsrfExemptSessionAuthentication])
def logout_view(request):
    logout(request)
    return Response({"detail": "Logged out"})


@api_view(["GET"])
def me_view(request):
    if not request.user.is_authenticated:
        return Response({"detail": "Not authenticated"}, status=401)

    # Enforce active blocking and trial checks for active sessions
    try:
        profile = None
        if hasattr(request.user, 'profile'):
            profile = request.user.profile
        else:
            profile, created = UserProfile.objects.get_or_create(user=request.user)
            if created:
                from datetime import date, timedelta
                profile.trial_expiry_date = date.today() + timedelta(days=90)
                profile.save()

        if profile:
            if profile.is_blocked:
                logout(request)
                return Response({"detail": "Your account has been blocked by the administrator."}, status=status.HTTP_403_FORBIDDEN)
            
            if not profile.is_portal_admin:
                company = profile.company
                if company and not company.is_upgraded:
                    from django.utils import timezone
                    from datetime import timedelta
                    expiry = profile.trial_expiry_date or (profile.created_at.date() + timedelta(days=90))
                    if timezone.now().date() > expiry:
                        logout(request)
                        return Response({
                            "detail": "Your 3-month trial period has expired. Please contact NMZ Associates to upgrade your account."
                        }, status=status.HTTP_403_FORBIDDEN)
    except Exception:
        pass

    company_name = "Default Business"
    role = "staff"
    owner_name = request.user.first_name or request.user.username
    is_portal_admin = False
    try:
        profile = None
        if hasattr(request.user, 'profile'):
            profile = request.user.profile
        else:
            profile, created = UserProfile.objects.get_or_create(user=request.user)
            if created:
                from datetime import date, timedelta
                profile.trial_expiry_date = date.today() + timedelta(days=90)
                profile.save()

        if profile:
            # Automatically designate 'admin' and 'moosa' accounts as portal admins (case-insensitive)
            if request.user.username.lower() in ['admin', 'moosa'] and not profile.is_portal_admin:
                profile.is_portal_admin = True
                profile.save()
                
            company_name = profile.company.name if profile.company else "Default Business"
            role = profile.role
            owner_name = profile.owner_name or request.user.first_name or request.user.username
            is_portal_admin = profile.is_portal_admin
    except Exception:
        pass
    company_logo = None
    try:
        if hasattr(request.user, 'profile') and request.user.profile.company:
            company_logo = request.user.profile.company.logo_base64
    except Exception:
        pass

    return Response({
        "username": request.user.username,
        "owner_name": owner_name,
        "is_staff": request.user.is_staff,
        "company_name": company_name,
        "role": role,
        "is_portal_admin": is_portal_admin,
        "company_logo": company_logo
    })


@api_view(["POST"])
def change_password_view(request):
    old_password = request.data.get("old_password")
    new_password = request.data.get("new_password")
    if not request.user.check_password(old_password):
        return Response({"detail": "Old password is incorrect."}, status=status.HTTP_400_BAD_REQUEST)
    try:
        validate_password(new_password, request.user)
    except ValidationError as error:
        return Response({"detail": error.messages}, status=status.HTTP_400_BAD_REQUEST)
    request.user.set_password(new_password)
    request.user.save()
    update_session_auth_hash(request, request.user)
    return Response({"detail": "Password changed."})


def period_summary(company, start, end):
    totals = Transaction.objects.filter(company=company, date__gte=start, date__lte=end).aggregate(
        income=Sum("amount", filter=Q(transaction_type="income")),
        expense=Sum("amount", filter=Q(transaction_type="expense")),
    )
    income = money(totals["income"])
    expense = money(totals["expense"])
    return {"income": income, "expense": expense, "balance": income - expense}


@api_view(["GET"])
def dashboard_view(request):
    today = timezone.localdate()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)
    year_start = today.replace(month=1, day=1)
    
    company = request.user.profile.company if request.user.is_authenticated and hasattr(request.user, 'profile') else None
    tx = Transaction.objects.filter(company=company)
    totals = tx.aggregate(
        income=Sum("amount", filter=Q(transaction_type="income")),
        expense=Sum("amount", filter=Q(transaction_type="expense")),
    )
    income = money(totals["income"])
    expense = money(totals["expense"])
    opening = money(Account.objects.filter(company=company).aggregate(total=Sum("opening_balance"))["total"])
    pending = DuePayment.objects.filter(company=company).exclude(status="paid")
    pending_totals = pending.aggregate(
        payable=Sum("amount", filter=Q(due_type="payable")),
        receivable=Sum("amount", filter=Q(due_type="receivable")),
    )
    
    accounts = Account.objects.filter(company=company).annotate(
        account_income=Sum("transactions__amount", filter=Q(transactions__transaction_type="income")),
        account_expense=Sum("transactions__amount", filter=Q(transactions__transaction_type="expense"))
    ).order_by("name")

    account_summaries = []
    for account in accounts:
        inc = money(account.account_income)
        exp = money(account.account_expense)
        account_summaries.append(
            {
                "id": account.id,
                "name": account.name,
                "account_type": account.account_type,
                "opening_balance": account.opening_balance,
                "income": inc,
                "expense": exp,
                "current_balance": account.opening_balance + inc - exp,
            }
        )

    return Response(
        {
            "totals": {
                "income": income,
                "expense": expense,
                "current_balance": opening + income - expense,
                "pending_payable": money(pending_totals["payable"]),
                "pending_receivable": money(pending_totals["receivable"]),
            },
            "periods": {
                "today": period_summary(company, today, today),
                "week": period_summary(company, week_start, today),
                "month": period_summary(company, month_start, today),
                "year": period_summary(company, year_start, today),
            },
            "account_summaries": account_summaries,
            "recent_transactions": TransactionSerializer(tx.select_related("category", "account", "party")[:8], many=True).data,
            "pending_dues": DuePaymentSerializer(pending.select_related("party")[:8], many=True).data,
        }
    )


@api_view(["GET"])
def reports_view(request):
    qs = filtered_transactions(request)
    by_date = (
        qs.annotate(grouped_date=TruncDate("date"))
        .values("grouped_date", "transaction_type")
        .annotate(total=Sum("amount"))
        .order_by("grouped_date")
    )
    totals = qs.aggregate(
        income=Sum("amount", filter=Q(transaction_type="income")),
        expense=Sum("amount", filter=Q(transaction_type="expense")),
    )
    income = money(totals["income"])
    expense = money(totals["expense"])
    return Response(
        {
            "summary": {"income": income, "expense": expense, "balance": income - expense},
            "by_date": [{"date": row["grouped_date"], "transaction_type": row["transaction_type"], "total": row["total"]} for row in by_date],
        }
    )


@api_view(["GET"])
def export_excel_view(request):
    qs = filtered_transactions(request)
    wb = Workbook()
    ws = wb.active
    ws.title = "Statement"
    ws.append(["Date", "Type", "Title", "Category", "Account", "Party", "Method", "Reference", "Debit", "Credit"])
    for tx in qs:
        ws.append([
            tx.date.isoformat(),
            tx.transaction_type,
            tx.title,
            tx.category.name,
            tx.account.name,
            tx.party.name if tx.party else "",
            tx.payment_method,
            tx.reference_number,
            tx.amount if tx.transaction_type == "expense" else "",
            tx.amount if tx.transaction_type == "income" else "",
        ])
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="account-statement.xlsx"'
    wb.save(response)
    return response


@api_view(["GET"])
def export_pdf_view(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="account-statement.pdf"'
    
    company = request.user.profile.company if request.user.is_authenticated and hasattr(request.user, 'profile') else None
    company_name = company.name if company else "LedgerPro"
    
    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    y = height - 55
    
    # Draw logo if exists
    if company and company.logo_base64:
        try:
            import base64
            from io import BytesIO
            from reportlab.lib.utils import ImageReader
            b64_data = company.logo_base64
            if ',' in b64_data:
                b64_data = b64_data.split(',', 1)[1]
            img_data = base64.b64decode(b64_data)
            img_file = BytesIO(img_data)
            reader = ImageReader(img_file)
            pdf.drawImage(reader, 40, y - 10, width=80, height=40, preserveAspectRatio=True, mask='auto')
            pdf.setFont("Helvetica-Bold", 16)
            pdf.drawString(130, y + 10, company_name)
            pdf.setFont("Helvetica", 10)
            pdf.drawString(130, y - 5, "Account Statement")
            y -= 50
        except Exception:
            pdf.setFont("Helvetica-Bold", 16)
            pdf.drawString(40, y, company_name)
            pdf.setFont("Helvetica", 10)
            pdf.drawString(40, y - 15, "Account Statement")
            y -= 40
    else:
        pdf.setFont("Helvetica-Bold", 16)
        pdf.drawString(40, y, company_name)
        pdf.setFont("Helvetica", 10)
        pdf.drawString(40, y - 15, "Account Statement")
        y -= 40
        
    pdf.setFont("Helvetica", 9)
    for tx in filtered_transactions(request):
        line = f"{tx.date} | {tx.transaction_type.upper()} | {tx.title} | {tx.category.name} | Rs {tx.amount}"
        pdf.drawString(40, y, line[:115])
        y -= 18
        if y < 50:
            pdf.showPage()
            y = height - 50
            pdf.setFont("Helvetica", 9)
    pdf.save()
    return response


class StockViewSet(viewsets.ModelViewSet):
    serializer_class = StockSerializer
    queryset = Stock.objects.all()

    def get_queryset(self):
        company = self.request.user.profile.company if self.request.user.is_authenticated and hasattr(self.request.user, 'profile') else None
        return Stock.objects.filter(company=company).annotate(
            sold_stock=Coalesce(Sum("sales__quantity"), Value(0)),
            remaining_stock=F("quantity") - Coalesce(Sum("sales__quantity"), Value(0))
        ).order_by("name")

    def perform_create(self, serializer):
        company = self.request.user.profile.company if self.request.user.is_authenticated and hasattr(self.request.user, 'profile') else None
        serializer.save(company=company)


class SaleViewSet(viewsets.ModelViewSet):
    serializer_class = SaleSerializer
    queryset = Sale.objects.select_related("stock", "account").order_by("-date", "-created_at")

    def get_queryset(self):
        company = self.request.user.profile.company if self.request.user.is_authenticated and hasattr(self.request.user, 'profile') else None
        return Sale.objects.filter(company=company).select_related("stock", "account").order_by("-date", "-created_at")

    def perform_create(self, serializer):
        company = self.request.user.profile.company if self.request.user.is_authenticated and hasattr(self.request.user, 'profile') else None
        serializer.save(company=company)


@api_view(["GET"])
def export_sales_excel_view(request):
    company = request.user.profile.company if request.user.is_authenticated and hasattr(request.user, 'profile') else None
    company_name = company.name if company else "LedgerPro"
    
    qs = Sale.objects.filter(company=company).select_related("stock", "account").order_by("-date", "-created_at")
    wb = Workbook()
    ws = wb.active
    ws.title = company_name[:30]
    try:
        ws.HeaderFooter.oddHeader.left.text = company_name
    except Exception:
        pass
    ws.append(["Date", "Stock Item", "Quantity", "Sale Price (Rs)", "Total Price (Rs)", "Account", "Notes", "Customer Name", "Customer Phone", "Customer Address"])
    for sale in qs:
        ws.append([
            sale.date.isoformat() if sale.date else "",
            sale.stock.name if sale.stock else "",
            sale.quantity,
            sale.sale_price,
            sale.total_price,
            sale.account.name if sale.account else "",
            sale.notes,
            sale.customer_name or "",
            sale.customer_phone or "",
            sale.customer_address or ""
        ])
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="sales-statement.xlsx"'
    wb.save(response)
    return response


@api_view(["GET"])
def export_sales_pdf_view(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="sales-statement.pdf"'
    
    company = request.user.profile.company if request.user.is_authenticated and hasattr(request.user, 'profile') else None
    company_name = company.name if company else "LedgerPro"
    
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor('#0f766e'),
        spaceAfter=4
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#64748b'),
        spaceAfter=15
    )
    
    cell_style = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#1e293b')
    )

    cell_header_style = ParagraphStyle(
        'CellHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white
    )

    # Decode base64 logo if exists
    logo_image = None
    if company and company.logo_base64:
        try:
            import base64
            from io import BytesIO
            from reportlab.platypus import Image
            b64_data = company.logo_base64
            if ',' in b64_data:
                b64_data = b64_data.split(',', 1)[1]
            img_data = base64.b64decode(b64_data)
            img_file = BytesIO(img_data)
            logo_image = Image(img_file, width=80, height=40)
        except Exception:
            pass

    header_table_data = []
    if logo_image:
        header_table_data.append([logo_image, Paragraph(company_name, title_style)])
    else:
        header_table_data.append([Paragraph(company_name, title_style)])
        
    header_table = Table(header_table_data, colWidths=[100, 400] if logo_image else [500])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
    ]))
    story.append(header_table)
    
    story.append(Paragraph("Sales Ledger Statement - Generated on " + timezone.now().strftime("%Y-%m-%d %H:%M"), subtitle_style))
    story.append(Spacer(1, 10))
    
    headers = [
        Paragraph("Date", cell_header_style),
        Paragraph("Stock Item", cell_header_style),
        Paragraph("Qty", cell_header_style),
        Paragraph("Rate (Rs)", cell_header_style),
        Paragraph("Total Price (Rs)", cell_header_style),
        Paragraph("Account", cell_header_style)
    ]
    
    table_data = [headers]
    
    company = request.user.profile.company if request.user.is_authenticated and hasattr(request.user, 'profile') else None
    sales = Sale.objects.filter(company=company).select_related("stock", "account").order_by("-date", "-created_at")
    total_sales_amount = Decimal(0)
    
    for sale in sales:
        total_sales_amount += sale.total_price or Decimal(0)
        table_data.append([
            Paragraph(sale.date.isoformat() if sale.date else "-", cell_style),
            Paragraph(sale.stock.name if sale.stock else "-", cell_style),
            Paragraph(str(sale.quantity), cell_style),
            Paragraph(f"Rs {sale.sale_price:,.0f}", cell_style),
            Paragraph(f"Rs {sale.total_price:,.0f}", cell_style),
            Paragraph(sale.account.name if sale.account else "-", cell_style)
        ])
        
    col_widths = [70, 150, 40, 75, 85, 95]
    sales_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    t_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f766e')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
    ])
    
    for i in range(1, len(table_data)):
        bg_color = colors.HexColor('#f8fafc') if i % 2 == 0 else colors.white
        t_style.add('BACKGROUND', (0, i), (-1, i), bg_color)
        
    sales_table.setStyle(t_style)
    story.append(sales_table)
    story.append(Spacer(1, 15))
    
    summary_data = [
        [Paragraph("<strong>Total Sales:</strong>", cell_style), Paragraph(f"<strong>Rs {total_sales_amount:,.0f}</strong>", cell_style)]
    ]
    summary_table = Table(summary_data, colWidths=[100, 150])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEBELOW', (0, 0), (-1, -1), 1.5, colors.HexColor('#0f766e')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    story.append(summary_table)
    doc.build(story)
    return response


@api_view(["POST"])
def import_transactions_view(request):
    company = request.user.profile.company if request.user.is_authenticated and hasattr(request.user, 'profile') else None
    if not company:
        return Response({"detail": "User has no associated company."}, status=status.HTTP_400_BAD_REQUEST)
        
    file = request.FILES.get("file")
    if not file:
        return Response({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)
        
    try:
        from datetime import datetime
        
        # Support PDF Import
        if file.name.lower().endswith(".pdf"):
            from pypdf import PdfReader
            import re
            
            reader = PdfReader(file)
            text = ""
            for page in reader.pages:
                text += page.extract_text() or ""
                
            pattern = re.compile(r'(\d{4}-\d{2}-\d{2})\s*\|\s*(INCOME|EXPENSE)\s*\|\s*([^|]+)\s*\|\s*([^|]+)\s*\|\s*(?:Rs\s*)?([\d,.]+)', re.IGNORECASE)
            
            count = 0
            from django.db import transaction as db_transaction
            
            with db_transaction.atomic():
                for line in text.splitlines():
                    match = pattern.search(line)
                    if not match:
                        continue
                        
                    raw_date, tx_type_str, title, cat_name, amt_str = match.groups()
                    tx_type = tx_type_str.lower()
                    title = title.strip()
                    cat_name = cat_name.strip()
                    amount = safe_decimal(amt_str)
                    tx_date = safe_date(raw_date)
                        
                    category, _ = Category.objects.get_or_create(
                        company=company,
                        name=cat_name,
                        category_type=tx_type,
                        defaults={"color": "#2563eb"}
                    )
                    
                    account, _ = Account.objects.get_or_create(
                        company=company,
                        name="Cash",
                        defaults={"account_type": "cash", "opening_balance": 0}
                    )
                    
                    Transaction.objects.create(
                        company=company,
                        transaction_type=tx_type,
                        title=title,
                        category=category,
                        account=account,
                        amount=amount,
                        date=tx_date,
                        payment_method="cash"
                    )
                    count += 1
                    
            return Response({"detail": f"Successfully imported {count} transactions from PDF."})
            
        # Default Excel Import
        from openpyxl import load_workbook
        wb = load_workbook(file, read_only=True)
        ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({"detail": "Empty Excel sheet."}, status=status.HTTP_400_BAD_REQUEST)
            
        header = [str(x).strip().lower() for x in rows[0] if x is not None]
        if "type" not in header or "category" not in header or "account" not in header:
            return Response({"detail": "Invalid Excel file format. Header must match the exported transaction statement format."}, status=status.HTTP_400_BAD_REQUEST)
            
        def idx_of(name):
            try:
                return header.index(name)
            except ValueError:
                return -1
                
        date_idx = idx_of("date")
        type_idx = idx_of("type")
        title_idx = idx_of("title")
        category_idx = idx_of("category")
        account_idx = idx_of("account")
        party_idx = idx_of("party")
        method_idx = idx_of("method")
        reference_idx = idx_of("reference")
        debit_idx = idx_of("debit")
        credit_idx = idx_of("credit")
        
        count = 0
        from django.db import transaction as db_transaction
        
        with db_transaction.atomic():
            for row in rows[1:]:
                if not any(row):
                    continue
                    
                tx_type = str(row[type_idx]).strip().lower() if type_idx != -1 and row[type_idx] is not None else "income"
                if tx_type not in ["income", "expense"]:
                    tx_type = "income"
                    
                title = str(row[title_idx]).strip() if title_idx != -1 and row[title_idx] is not None else "Imported Transaction"
                
                raw_date = row[date_idx] if date_idx != -1 else None
                tx_date = safe_date(raw_date)
                    
                debit_val = row[debit_idx] if debit_idx != -1 else None
                credit_val = row[credit_idx] if credit_idx != -1 else None
                
                if tx_type == "expense":
                    amount = safe_decimal(debit_val)
                else:
                    amount = safe_decimal(credit_val)
                    
                if amount <= 0:
                    amount = safe_decimal(credit_val or debit_val)
                    
                cat_name = str(row[category_idx]).strip() if category_idx != -1 and row[category_idx] is not None else "Uncategorized"
                category, _ = Category.objects.get_or_create(
                    company=company,
                    name=cat_name,
                    category_type=tx_type,
                    defaults={"color": "#2563eb"}
                )
                
                acc_name = str(row[account_idx]).strip() if account_idx != -1 and row[account_idx] is not None else "Cash"
                account, _ = Account.objects.get_or_create(
                    company=company,
                    name=acc_name,
                    defaults={"account_type": "cash", "opening_balance": 0}
                )
                
                party = None
                p_name = str(row[party_idx]).strip() if party_idx != -1 and row[party_idx] is not None else ""
                if p_name:
                    party, _ = Party.objects.get_or_create(
                        company=company,
                        name=p_name,
                        defaults={"party_type": "customer"}
                    )
                    
                method = str(row[method_idx]).strip().lower() if method_idx != -1 and row[method_idx] is not None else "cash"
                ref_num = str(row[reference_idx]).strip() if reference_idx != -1 and row[reference_idx] is not None else ""
                
                Transaction.objects.create(
                    company=company,
                    transaction_type=tx_type,
                    title=title,
                    category=category,
                    account=account,
                    party=party,
                    amount=amount,
                    date=tx_date,
                    payment_method=method,
                    reference_number=ref_num
                )
                count += 1
                
        return Response({"detail": f"Successfully imported {count} transactions."})
    except Exception as e:
        return Response({"detail": f"Error parsing file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
def import_sales_view(request):
    company = request.user.profile.company if request.user.is_authenticated and hasattr(request.user, 'profile') else None
    if not company:
        return Response({"detail": "User has no associated company."}, status=status.HTTP_400_BAD_REQUEST)
        
    file = request.FILES.get("file")
    if not file:
        return Response({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)
        
    try:
        from datetime import datetime
        
        # Support PDF Import
        if file.name.lower().endswith(".pdf"):
            from pypdf import PdfReader
            import re
            
            reader = PdfReader(file)
            text = ""
            for page in reader.pages:
                text += page.extract_text() or ""
                
            # Regex for PDF Sales ledger table rows
            pattern = re.compile(r'^\s*(\d{4}-\d{2}-\d{2})\s+(.+?)\s+(\d+)\s+(?:Rs\s*)?([\d,.]+)\s+(?:Rs\s*)?([\d,.]+)\s+(.+?)\s*$', re.MULTILINE)
            
            count = 0
            from django.db import transaction as db_transaction
            
            with db_transaction.atomic():
                for match in pattern.finditer(text):
                    raw_date, stock_name, qty_str, price_str, total_str, acc_name = match.groups()
                    
                    stock_name = stock_name.strip()
                    qty = safe_int(qty_str)
                    price = safe_decimal(price_str)
                    acc_name = acc_name.strip()
                    sale_date = safe_date(raw_date)
                        
                    stock, _ = Stock.objects.get_or_create(
                        company=company,
                        name=stock_name,
                        defaults={"quantity": 10000, "unit_price": 0}
                    )
                    
                    account, _ = Account.objects.get_or_create(
                        company=company,
                        name=acc_name,
                        defaults={"account_type": "cash", "opening_balance": 0}
                    )
                    
                    Sale.objects.create(
                        company=company,
                        stock=stock,
                        quantity=qty,
                        sale_price=price,
                        account=account,
                        date=sale_date,
                        notes="Imported from PDF."
                    )
                    count += 1
                    
            return Response({"detail": f"Successfully imported {count} sales from PDF."})
            
        # Default Excel Import
        from openpyxl import load_workbook
        wb = load_workbook(file, read_only=True)
        ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({"detail": "Empty Excel sheet."}, status=status.HTTP_400_BAD_REQUEST)
            
        header = [str(x).strip().lower() for x in rows[0] if x is not None]
        if "stock item" not in header or "quantity" not in header or "account" not in header:
            return Response({"detail": "Invalid Excel file format. Header must match the exported sales statement format."}, status=status.HTTP_400_BAD_REQUEST)
            
        def idx_of(name):
            try:
                return header.index(name)
            except ValueError:
                return -1
                
        date_idx = idx_of("date")
        stock_idx = idx_of("stock item")
        quantity_idx = idx_of("quantity")
        price_idx = idx_of("sale price (rs)")
        account_idx = idx_of("account")
        notes_idx = idx_of("notes")
        cust_name_idx = idx_of("customer name")
        cust_phone_idx = idx_of("customer phone")
        cust_addr_idx = idx_of("customer address")
        
        count = 0
        from django.db import transaction as db_transaction
        
        with db_transaction.atomic():
            for row in rows[1:]:
                if not any(row):
                    continue
                    
                stock_name = str(row[stock_idx]).strip() if stock_idx != -1 and row[stock_idx] is not None else "Imported Item"
                stock, _ = Stock.objects.get_or_create(
                    company=company,
                    name=stock_name,
                    defaults={"quantity": 10000, "unit_price": 0}
                )
                
                qty = safe_int(row[quantity_idx]) if quantity_idx != -1 else 1
                price = safe_decimal(row[price_idx]) if price_idx != -1 else Decimal(0)
                
                acc_name = str(row[account_idx]).strip() if account_idx != -1 and row[account_idx] is not None else "Cash"
                account, _ = Account.objects.get_or_create(
                    company=company,
                    name=acc_name,
                    defaults={"account_type": "cash", "opening_balance": 0}
                )
                
                raw_date = row[date_idx] if date_idx != -1 else None
                sale_date = safe_date(raw_date)
                    
                notes = str(row[notes_idx]).strip() if notes_idx != -1 and row[notes_idx] is not None else ""
                cust_name = str(row[cust_name_idx]).strip() if cust_name_idx != -1 and row[cust_name_idx] is not None else ""
                cust_phone = str(row[cust_phone_idx]).strip() if cust_phone_idx != -1 and row[cust_phone_idx] is not None else ""
                cust_addr = str(row[cust_addr_idx]).strip() if cust_addr_idx != -1 and row[cust_addr_idx] is not None else ""
                
                Sale.objects.create(
                    company=company,
                    stock=stock,
                    quantity=qty,
                    sale_price=price,
                    account=account,
                    date=sale_date,
                    notes=notes,
                    customer_name=cust_name,
                    customer_phone=cust_phone,
                    customer_address=cust_addr
                )
                count += 1
                
        return Response({"detail": f"Successfully imported {count} sales."})
    except Exception as e:
        return Response({"detail": f"Error parsing file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)


def safe_int(val, default=0):
    if val is None:
        return default
    try:
        clean_val = str(val).replace(',', '').strip()
        if not clean_val:
            return default
        return int(float(clean_val))
    except (ValueError, TypeError):
        return default


def safe_decimal(val, default=Decimal('0')):
    if val is None:
        return default
    try:
        clean_val = str(val).replace('Rs', '').replace('rs', '').replace(',', '').strip()
        if not clean_val:
            return default
        return Decimal(clean_val)
    except (ValueError, TypeError, ArithmeticError):
        return default


def safe_date(raw_date):
    from datetime import datetime, date
    if isinstance(raw_date, (datetime, date)):
        return raw_date.date() if isinstance(raw_date, datetime) else raw_date
    if isinstance(raw_date, str):
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(raw_date.strip(), fmt).date()
            except ValueError:
                continue
    return timezone.localdate()


@api_view(["GET"])
def export_stock_excel_view(request):
    company = request.user.profile.company if request.user.is_authenticated and hasattr(request.user, 'profile') else None
    company_name = company.name if company else "LedgerPro"
    
    stock_qs = Stock.objects.filter(company=company).annotate(
        sold_stock=Coalesce(Sum("sales__quantity"), Value(0)),
        remaining_stock=F("quantity") - Coalesce(Sum("sales__quantity"), Value(0))
    ).order_by("name")
    
    wb = Workbook()
    ws = wb.active
    ws.title = company_name[:30]
    try:
        ws.HeaderFooter.oddHeader.left.text = company_name
    except Exception:
        pass
    ws.append(["Item Name", "Unit Price (Rs)", "Total Stock", "Sold Stock", "Remaining Stock", "Value (Remaining)"])
    
    for item in stock_qs:
        rem_stock = item.remaining_stock if item.remaining_stock is not None else 0
        price = item.unit_price if item.unit_price is not None else Decimal(0)
        ws.append([
            item.name,
            price,
            item.quantity,
            item.sold_stock or 0,
            rem_stock,
            rem_stock * price
        ])
        
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="stock-inventory.xlsx"'
    wb.save(response)
    return response


@api_view(["GET"])
def export_stock_pdf_view(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="stock-inventory.pdf"'
    
    company = request.user.profile.company if request.user.is_authenticated and hasattr(request.user, 'profile') else None
    company_name = company.name if company else "LedgerPro"
    
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor('#0f766e'),
        spaceAfter=4
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#64748b'),
        spaceAfter=15
    )
    
    cell_style = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#1e293b')
    )

    cell_header_style = ParagraphStyle(
        'CellHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white
    )

    # Decode base64 logo if exists
    logo_image = None
    if company and company.logo_base64:
        try:
            import base64
            from io import BytesIO
            from reportlab.platypus import Image
            b64_data = company.logo_base64
            if ',' in b64_data:
                b64_data = b64_data.split(',', 1)[1]
            img_data = base64.b64decode(b64_data)
            img_file = BytesIO(img_data)
            logo_image = Image(img_file, width=80, height=40)
        except Exception:
            pass

    header_table_data = []
    if logo_image:
        header_table_data.append([logo_image, Paragraph(company_name, title_style)])
    else:
        header_table_data.append([Paragraph(company_name, title_style)])
        
    header_table = Table(header_table_data, colWidths=[100, 400] if logo_image else [500])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
    ]))
    story.append(header_table)
    
    story.append(Paragraph("Stock Inventory Statement - Generated on " + timezone.now().strftime("%Y-%m-%d %H:%M"), subtitle_style))
    story.append(Spacer(1, 10))
    
    headers = [
        Paragraph("Item Name", cell_header_style),
        Paragraph("Unit Price (Rs)", cell_header_style),
        Paragraph("Total Stock", cell_header_style),
        Paragraph("Sold Stock", cell_header_style),
        Paragraph("Remaining Stock", cell_header_style),
        Paragraph("Value (Remaining)", cell_header_style)
    ]
    
    table_data = [headers]
    
    company = request.user.profile.company if request.user.is_authenticated and hasattr(request.user, 'profile') else None
    stock_qs = Stock.objects.filter(company=company).annotate(
        sold_stock=Coalesce(Sum("sales__quantity"), Value(0)),
        remaining_stock=F("quantity") - Coalesce(Sum("sales__quantity"), Value(0))
    ).order_by("name")
    
    total_val = Decimal(0)
    for item in stock_qs:
        rem_stock = item.remaining_stock if item.remaining_stock is not None else 0
        price = item.unit_price if item.unit_price is not None else Decimal(0)
        val = rem_stock * price
        total_val += val
        table_data.append([
            Paragraph(item.name, cell_style),
            Paragraph(f"Rs {price:,.0f}", cell_style),
            Paragraph(str(item.quantity or 0), cell_style),
            Paragraph(str(item.sold_stock or 0), cell_style),
            Paragraph(str(rem_stock), cell_style),
            Paragraph(f"Rs {val:,.0f}", cell_style)
        ])
        
    col_widths = [135, 80, 65, 65, 70, 100]
    stock_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    t_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f766e')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
    ])
    
    for i in range(1, len(table_data)):
        bg_color = colors.HexColor('#f8fafc') if i % 2 == 0 else colors.white
        t_style.add('BACKGROUND', (0, i), (-1, i), bg_color)
        
    stock_table.setStyle(t_style)
    story.append(stock_table)
    story.append(Spacer(1, 15))
    
    summary_data = [
        [Paragraph("<strong>Total Inventory Value:</strong>", cell_style), Paragraph(f"<strong>Rs {total_val:,.0f}</strong>", cell_style)]
    ]
    summary_table = Table(summary_data, colWidths=[150, 150])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEBELOW', (0, 0), (-1, -1), 1.5, colors.HexColor('#0f766e')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    story.append(summary_table)
    doc.build(story)
    return response


@api_view(["POST"])
def import_stock_view(request):
    company = request.user.profile.company if request.user.is_authenticated and hasattr(request.user, 'profile') else None
    if not company:
        return Response({"detail": "User has no associated company."}, status=status.HTTP_400_BAD_REQUEST)
        
    file = request.FILES.get("file")
    if not file:
        return Response({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)
        
    try:
        count = 0
        from django.db import transaction as db_transaction
        
        # Support PDF Import
        if file.name.lower().endswith(".pdf"):
            from pypdf import PdfReader
            import re
            
            reader = PdfReader(file)
            text = ""
            for page in reader.pages:
                text += page.extract_text() or ""
                
            pattern = re.compile(r'^\s*(.+?)\s+(?:Rs\s*)?([\d,.]+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(?:Rs\s*)?([\d,.]+)\s*$', re.MULTILINE)
            
            with db_transaction.atomic():
                for match in pattern.finditer(text):
                    name, price_str, total_str, sold_str, rem_str, val_str = match.groups()
                    name = name.strip()
                    price = safe_decimal(price_str)
                    total = safe_int(total_str)
                    
                    stock, created = Stock.objects.get_or_create(
                        company=company,
                        name=name,
                        defaults={"quantity": total, "unit_price": price}
                    )
                    if not created:
                        stock.quantity = total
                        stock.unit_price = price
                        stock.save()
                    count += 1
                    
            return Response({"detail": f"Successfully imported/updated {count} stock items from PDF."})
            
        # Default Excel Import
        from openpyxl import load_workbook
        wb = load_workbook(file, read_only=True)
        ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({"detail": "Empty Excel sheet."}, status=status.HTTP_400_BAD_REQUEST)
            
        header = [str(x).strip().lower() for x in rows[0] if x is not None]
        if "item name" not in header or "total stock" not in header or "unit price (rs)" not in header:
            return Response({"detail": "Invalid Excel file format. Header must match the exported stock inventory format."}, status=status.HTTP_400_BAD_REQUEST)
            
        def idx_of(name):
            try:
                return header.index(name)
            except ValueError:
                return -1
                
        name_idx = idx_of("item name")
        total_idx = idx_of("total stock")
        price_idx = idx_of("unit price (rs)")
        
        with db_transaction.atomic():
            for row in rows[1:]:
                if not any(row):
                    continue
                    
                name = str(row[name_idx]).strip() if name_idx != -1 and row[name_idx] is not None else ""
                if not name:
                    continue
                    
                total = safe_int(row[total_idx])
                price = safe_decimal(row[price_idx])
                
                stock, created = Stock.objects.get_or_create(
                    company=company,
                    name=name,
                    defaults={"quantity": total, "unit_price": price}
                )
                if not created:
                    stock.quantity = total
                    stock.unit_price = price
                    stock.save()
                count += 1
                
        return Response({"detail": f"Successfully imported/updated {count} stock items."})
    except Exception as e:
        return Response({"detail": f"Error parsing file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def superadmin_users_view(request):
    try:
        profile = request.user.profile
    except Exception:
        return Response({"detail": "Profile not found."}, status=403)
        
    if not profile.is_portal_admin:
        return Response({"detail": "Access denied."}, status=403)
        
    # Get all profiles
    profiles = UserProfile.objects.select_related("user", "company").order_by("-created_at")
    data = []
    from django.utils import timezone
    from datetime import timedelta
    
    for p in profiles:
        # Enforce that ONLY 'admin' and 'moosa' accounts are portal admins
        should_be_admin = p.user.username.lower() in ['admin', 'moosa']
        if p.is_portal_admin != should_be_admin:
            p.is_portal_admin = should_be_admin
            p.save()

        company = p.company
        is_upgraded = company.is_upgraded if company else False
        
        # Determine trial status
        is_expired = False
        expiry = p.trial_expiry_date or (p.created_at.date() + timedelta(days=90))
        if not p.is_portal_admin and company and not is_upgraded:
            is_expired = timezone.now().date() > expiry
            
        data.append({
            "id": p.id,
            "username": p.user.username,
            "owner_name": p.owner_name or p.user.first_name or p.user.username,
            "business_name": p.business_name or (company.name if company else ""),
            "phone": p.phone,
            "joining_date": p.created_at.isoformat(),
            "trial_expiry_date": p.trial_expiry_date.isoformat() if p.trial_expiry_date else expiry.isoformat(),
            "is_upgraded": is_upgraded,
            "is_blocked": p.is_blocked,
            "is_expired": is_expired,
            "is_portal_admin": p.is_portal_admin
        })
        
    return Response(data)

@api_view(["POST"])
@authentication_classes([CsrfExemptSessionAuthentication])
@permission_classes([IsAuthenticated])
def superadmin_toggle_block_view(request, profile_id):
    try:
        profile = request.user.profile
    except Exception:
        return Response({"detail": "Profile not found."}, status=403)
        
    if not profile.is_portal_admin:
        return Response({"detail": "Access denied."}, status=403)
        
    try:
        target_profile = UserProfile.objects.get(id=profile_id)
        if target_profile.is_portal_admin:
            return Response({"detail": "Cannot block the portal superadmin."}, status=400)
            
        target_profile.is_blocked = not target_profile.is_blocked
        target_profile.save()
        
        return Response({
            "success": True,
            "is_blocked": target_profile.is_blocked,
            "message": f"User {'blocked' if target_profile.is_blocked else 'unblocked'} successfully."
        })
    except UserProfile.DoesNotExist:
        return Response({"detail": "User not found."}, status=404)

@api_view(["POST"])
@authentication_classes([CsrfExemptSessionAuthentication])
@permission_classes([IsAuthenticated])
def superadmin_toggle_upgrade_view(request, profile_id):
    try:
        profile = request.user.profile
    except Exception:
        return Response({"detail": "Profile not found."}, status=403)
        
    if not profile.is_portal_admin:
        return Response({"detail": "Access denied."}, status=403)
        
    try:
        target_profile = UserProfile.objects.get(id=profile_id)
        company = target_profile.company
        if not company:
            return Response({"detail": "No company associated with this user."}, status=400)
            
        company.is_upgraded = not company.is_upgraded
        company.save()
        
        return Response({
            "success": True,
            "is_upgraded": company.is_upgraded,
            "message": f"Company plan {'upgraded to Premium' if company.is_upgraded else 'downgraded to Free Trial'} successfully."
        })
    except UserProfile.DoesNotExist:
        return Response({"detail": "User not found."}, status=404)


@api_view(["POST"])
@authentication_classes([CsrfExemptSessionAuthentication])
@permission_classes([IsAuthenticated])
def superadmin_set_expiry_view(request, profile_id):
    try:
        profile = request.user.profile
    except Exception:
        return Response({"detail": "Profile not found."}, status=403)
        
    if not profile.is_portal_admin:
        return Response({"detail": "Access denied."}, status=403)
        
    expiry_date_str = request.data.get("expiry_date")
    if not expiry_date_str:
        return Response({"detail": "Missing expiry_date field."}, status=400)
        
    try:
        from datetime import datetime
        expiry_date = datetime.strptime(expiry_date_str, "%Y-%m-%d").date()
    except ValueError:
        return Response({"detail": "Invalid date format. Use YYYY-MM-DD."}, status=400)
        
    try:
        target_profile = UserProfile.objects.get(id=profile_id)
        if target_profile.is_portal_admin:
            return Response({"detail": "Cannot modify superadmin trial expiry."}, status=400)
            
        target_profile.trial_expiry_date = expiry_date
        target_profile.save()
        
        return Response({
            "success": True,
            "trial_expiry_date": target_profile.trial_expiry_date.isoformat(),
            "message": f"Trial expiry date updated to {target_profile.trial_expiry_date} successfully."
        })
    except UserProfile.DoesNotExist:
        return Response({"detail": "User not found."}, status=404)


@api_view(["GET"])
@permission_classes([AllowAny])
def test_error_view(request):
    import traceback
    try:
        from accounts.models import Transaction, Category, Account, Party
        from accounts.serializers import TransactionSerializer
        from datetime import date
        
        category = Category.objects.get(id=9)
        account = Account.objects.get(id=12)
        party = Party.objects.get(id=11)
        
        from django.db import transaction as db_transaction
        with db_transaction.atomic():
            # Test 1: With Party
            tx_with_party = Transaction.objects.create(
                company=category.company,
                transaction_type="income",
                title="Mock Transaction Save test with Party",
                category=category,
                account=account,
                party=party,
                amount=100.00,
                date=date.today(),
                payment_method="cash"
            )
            serializer_with_party = TransactionSerializer(tx_with_party)
            data_with_party = serializer_with_party.data

            # Test 2: Without Party
            tx_no_party = Transaction.objects.create(
                company=category.company,
                transaction_type="income",
                title="Mock Transaction Save test no Party",
                category=category,
                account=account,
                party=None,
                amount=100.00,
                date=date.today(),
                payment_method="cash"
            )
            serializer_no_party = TransactionSerializer(tx_no_party)
            data_no_party = serializer_no_party.data
            
            db_transaction.set_rollback(True)
            
        return Response({
            "success": True,
            "detail": "Test transactions serialized successfully!",
            "data_with_party": data_with_party,
            "data_no_party": data_no_party
        })
    except Exception as e:
        return Response({
            "error": str(e),
            "traceback": traceback.format_exc()
        }, status=500)


